import openpyxl
from openpyxl.styles import Border, Side
import shutil
import os
from .hazop_structure import *


hazop_template_path = "data\\HAZOP양식.xlsx"
hazop_template_worksheet_name = "HAZOP양식"

worksheet_type = openpyxl.workbook.workbook.Worksheet


def __merge_cells(ws: worksheet_type, start_cell_index, end_cell_index):
    ws.merge_cells(start_row=start_cell_index[0], start_column=start_cell_index[1],
                   end_row=end_cell_index[0], end_column=end_cell_index[1])


def __set_value(ws: worksheet_type, cell_index, str_data: str):
    ws.cell(row=cell_index[0], column=cell_index[1]).value = str_data


def __copy_sheet(root_node: tree_node, wb: openpyxl.Workbook):
    sample_ws = wb[hazop_template_worksheet_name]
    new_ws = wb.copy_worksheet(sample_ws)
    if len(root_node.data) > 0:
        new_ws.title = root_node.data
        wb.move_sheet(root_node.data, -3)
    else:
        new_ws.title = " "
        wb.move_sheet(" ", -3)

    return new_ws


def __generate_main_text(ws, tree_node_: tree_node, row, column):
    cell_str = tree_node_.data
    __set_value(ws, (row, column), cell_str)
    __merge_cells(ws, (row, column), (row + tree_node_.get_row_count() - 1, column))
    for child_node in tree_node_.child_list:
        __generate_main_text(ws, child_node, row, column+1)
        row += child_node.get_row_count()


def line_borders(ws, start_index, height, width):
    box_border = Border(left=Side(border_style="thin",
                                  color='000000'),
                        right=Side(border_style="thin",
                                   color='000000'),
                        top=Side(border_style="thin",
                                 color='000000'),
                        bottom=Side(border_style="thin",
                                    color='000000'))

    for row_count, row in enumerate(ws.rows):
        if row_count >= start_index[0] - 1:
            for col_count, cell in enumerate(row):
                if col_count >= start_index[1] - 1:
                    cell.border = box_border


def generate_from_data(root_node: tree_node, save_path):
    header_length = 20
    shutil.copy(os.path.abspath(hazop_template_path), os.path.abspath(save_path))
    pm_node = root_node.child_list[0]
    if save_path != ".xlsx":
        wb = openpyxl.load_workbook(filename=save_path)
        ws = __copy_sheet(pm_node, wb)
        row, column = 7, 2
        __generate_main_text(ws, pm_node, row, column)

        line_borders(ws, (row, column), root_node.get_row_count(), header_length)

        template_worksheet = wb.get_sheet_by_name(hazop_template_worksheet_name)
        wb.remove(template_worksheet)
        wb.save(save_path)
        wb.close()

