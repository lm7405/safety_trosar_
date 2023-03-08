import io

import pandas
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtCore import Qt, QEvent
from pandas import DataFrame
from typing import List, Tuple, Any, Dict, Union, ClassVar
from pydantic import BaseModel
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.merge import CellRange
from openpyxl.utils.cell import coordinate_from_string  # https://www.crocus.co.kr/1759

temp_column_number = 9


def index_to_cell_letter(row_, column_):
    string = get_column_letter(column_ + 1) + str(row_ + 1)
    return string


# 'B3'을 (2, 1)로 변환, 에러 시 None 반환
def index_from_string(cell_string: str):
    row = 0
    column = 0

    number_flag = False
    for letter in cell_string:
        if letter.isupper():
            if number_flag is True:
                return None
            column = column * 26 + ord(letter) - ord('A')
        elif letter.isdigit():
            row = row * 10 + int(letter)
            number_flag = True
        else:
            return None
    row -= 1

    return row, column


# (1, 2)를 'C2'로 변환
def string_from_index(index: Tuple[int, int]):
    row = index[0]
    column = index[1]

    return index_to_cell_letter(row, column)


# 데이터의 원본을 담당함
# data_list와 dataframe과의 호환을 보장하고자 함        data_list가 왜 필요하지? 필요없어보이는데.
# 뷰는 excel로 관리되며, 데이터는 dataframe으로 관리됨
# dataframe과 내부 변수를 사용하여 뷰를 생성할 수 있어야 함. 역도 마찬가지
# 엑셀과 내부변수로 dataframe을, dataframe과 내부변수로 엑셀을 만들 수 있어야 함
# 상세 생동은 아래 변수에 대한 주석을 참조
class PandasTableData(BaseModel):           # 데이터프레임에 대한 변수

    column_labels: Dict[str, str]
    cell_offset: Tuple[int, int] = (0, 0)
    merged_cell_ranges: List[Union[CellRange]] = []
    protected_cell_ranges: List[Union[CellRange]] = []
    dataframe: Union[DataFrame, None]
    column_labels: ClassVar[List[Tuple[str, str]]]

    NON_AVAILABLE_TEXT_LIST = ["N/A", "-", "x", "X"]

    # # # #      데이터 관리를 위한 기본 함수들

    # base model의 import 격차를 위해 사용
    def __init__(self, **data: Any):
        if isinstance(data, dict):
            if "merged_cell_ranges" in data:
                target = data["merged_cell_ranges"]
                tmp = []
                if len(target) > 0 and isinstance(target[-1], str):
                    for item in target:
                        tmp.append(CellRange(item))
                    data["merged_cell_ranges"] = tmp
            if "protected_cell_ranges" in data:
                target = data["protected_cell_ranges"]
                tmp = []
                if len(target) > 0 and isinstance(target[-1], str):
                    for item in target:
                        tmp.append(CellRange(item))
                    data["protected_cell_ranges"] = tmp

        if isinstance(data, dict):
            if "dataframe" in data and data["dataframe"] is not None:
                buffer = io.StringIO()
                buffer.write(data["dataframe"])
                buffer.seek(0)
                data["dataframe"] = pandas.read_csv(buffer, keep_default_na=False)

        super().__init__(**data)

    class Config:
        arbitrary_types_allowed = True

    @classmethod
    def column_name(cls, key: str):
        found = [i for i in cls.column_labels if i[0] == key]
        return found[0][1]

    def init_dataframe(self):
        header = self.get_header()
        empty_dict = {}
        for key in header:
            empty_dict[key] = []
        self.set_dataframe(DataFrame(empty_dict))

    def set_dataframe(self, dataframe: DataFrame):
        header = self.get_header()
        new_header = dataframe.columns
        for i, column_name in enumerate(header):
            if column_name != new_header[i]:
                raise Exception("At 'set_dataframe', Header is something wrong")
        self.dataframe = dataframe

    def set_cell_protected(self, cell_range: CellRange):
        self.protected_cell_ranges.append(cell_range)

    def set_cell_unprotected(self, cell_range: CellRange):
        raise

    # Export 격차를 위해 사용
    def dict(
        self,
        *,
        include=None,
        exclude=None,
        by_alias: bool = False,
        skip_defaults: bool = None,
        exclude_unset: bool = False,
        exclude_defaults: bool = False,
        exclude_none: bool = False,
    ):
        if self.dataframe is not None:
            output_: Dict[str, Any] = {
                "merged_cell_ranges": [],
                "protected_cell_ranges": [],
                "dataframe": self.dataframe.to_csv(index=False),
                "column_labels": self.column_labels
            }
        else:
            output_: Dict[str, Any] = {
                "merged_cell_ranges": [],
                "protected_cell_ranges": [],
                "dataframe": None,
                "column_labels": self.column_labels
            }
        for merged_cell_range in self.merged_cell_ranges:
            output_["merged_cell_ranges"].append(str(merged_cell_range))
        for protected_cell_range in self.protected_cell_ranges:
            output_["protected_cell_ranges"].append(str(protected_cell_range))

        return output_

    def save_to_excel(self, path):
        from io import BytesIO
        file = BytesIO()
        self.dataframe.to_excel(file, sheet_name='Sheet1', columns=None, header=True,
                                index=True, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True,
                                inf_rep='inf', freeze_panes=None,  # 행, 열 고정
                                storage_options=None)
        file.seek(0)
        with open(path, "wb") as f:
            f.write(file.read())
        return file

    def import_dataframe_row_from_obj(self):
        raise

    def get_header(self):
        columns = [i[0] for i in self.column_labels]
        return columns

    def insert_row(self, row: List):
        # add line to dataframe
        # add line to tableView
        # modify merged setting
        # modify protected setting

        raise

    def delete_row(self, row: List):
        # check if deletable
        # delete line in dataframe
        # delete line in tableView
        # modify merged setting
        # modify protected setting

        raise

    def undo(self):
        raise

    def redo(self):
        raise

    def copy_data(self):
        raise

    def paste_data(self, paste_data):
        raise

    @staticmethod
    def index_to_cell_range(start_row, start_column, end_row, end_column):
        start_cell = index_to_cell_letter(start_row, start_column)
        end_cell = index_to_cell_letter(end_row, end_column)
        range_string = start_cell + ':' + end_cell
        return CellRange(range_string)

    @staticmethod
    def index_to_cell_letter(row_, column_):
        string = get_column_letter(column_ + 1) + str(row_ + 1)
        return string

    @staticmethod
    def is_cell_range_contain(cell_text: str, cell_range_list: List[CellRange]):
        for cell_range in cell_range_list:
            if cell_text in cell_range:
                return True
        return False

    def get_value(self, row: int, column: int):
        return self.dataframe.iat[row][column]

    def set_value(self, row: int, column: int, value: str):
        self.dataframe.iat[row][column] = value

    def get_row(self, row):
        return self.dataframe[row]

    @staticmethod
    def get_shown_row(row_index) -> str:
        return str(row_index)

    def get_shown_column(self, column_index) -> str:
        if column_index >= len(self.get_header()):
            print()
        return self.column_name(self.get_header()[column_index])

    def get_shown_header(self):
        return [self.column_name(i) for i in self.get_header()]


# QTableView 제어에 필요한 공용 펑션을 정의
class PandasTableViewMgr:
    table_data: PandasTableData
    pandas_model: QtCore.QAbstractTableModel
    qtableview: QtWidgets.QTableView

    def __init__(self, qtableview: QtWidgets.QTableView, model_parant=None):
        self.qtableview = qtableview
        self.pandas_model = PandasModel(model_parant)

        """
        at widget(pandasTableMgr),
        for row in range(0, self._tm.rowCount()):
            self._tv.openPersistentEditor(self._tm.index(row, 4))
        """

    # dataframe 초기화         # parent=gui widget
    def init_data(self, table_data: PandasTableData,
                  checkbox_delegate: List[int] = None,
                  combobox_delegate: List[Tuple[int, List[str]]] = None):
        self.table_data = table_data
        self.pandas_model.init_data(self.table_data)
        self.qtableview.setModel(self.pandas_model)

        # Checkbox
        if checkbox_delegate is not None:
            for column in checkbox_delegate:
                self.qtableview.setItemDelegateForColumn(column, CheckBoxDelegate(self.qtableview))

        if combobox_delegate is not None:
            for column, data in combobox_delegate:
                # Set Delegate
                self.qtableview.setItemDelegateForColumn(
                    column, ComboDelegate(self.qtableview, data)
                )

                # openPersistentEditor
                for row_count, _ in self.table_data.dataframe.iterrows():
                    if isinstance(row_count, int):
                        # self.qtableview.openPersistentEditor(self.pandas_model.index(row_count, column))
                        pass
                    else:
                        raise
        self.show_table()

    # qtableview 반영
    def show_table(self):
        self.merge_table()
        self.qtableview.show()

    # 적용 대상 qtableview가 변경될 경우
    def change_qtableview(self, qtableview: QtWidgets.QTableView):
        self.qtableview = qtableview
        self.show_table()

    # 셀 병합 정보 적용
    def merge_table(self):
        row_offset = 0
        column_offset = 0
        if self.table_data.cell_offset is not None:
            row_offset = self.table_data.cell_offset[0]
            column_offset = self.table_data.cell_offset[1]

        if self.table_data.merged_cell_ranges is not None:
            for merged_cell_range in self.table_data.merged_cell_ranges:
                start_row = merged_cell_range.min_row - 1
                start_column = merged_cell_range.min_col - 1
                span_size = merged_cell_range.size

                rowSpanCount = span_size["rows"]
                columnSpanCount = span_size["columns"]

                # 아래 코드 확인 필요(헤더가 merge range 를 침범할 경우의 조건문)
                if (start_row - row_offset < 0) or (start_column - column_offset < 0):
                    pass
                else:
                    self.qtableview.setSpan(start_row - row_offset, start_column - column_offset,
                                            rowSpanCount, columnSpanCount)

    # 셀 병합 확인
    def is_cell_merged(self, cell_index):
        # cell = sheet.cell(row, column)
        # for mergedCell in self.merged_cell_ranges:
        #     if cell.coordinate in mergedCell:
        #         return True
        for mergedCell in self.table_data.merged_cell_ranges:
            if string_from_index(cell_index) in mergedCell:
                return True
        return False

    def insert_row(self):
        cell_index = 1
        # add line to dataframe
        # add line to tableView
        # modify merged setting
        # modify protected setting
        raise

    def delete_row(self):
        # check if deletable
        # delete line in dataframe
        # delete line in tableView
        # modify merged setting
        # modify protected setting
        raise


# pandas dataframe 을 qtableview 의 model 로 사용하기 위한 클래스
# QAbstractTableModel 이 갖고 있는 기능들을 dataframe 과 연동되도록 재정의 하는 것이 목적
class PandasModel(QtCore.QAbstractTableModel):
    __flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable
    __event_flag: bool = False
    table_data: PandasTableData

    def __init__(self, table_data: PandasTableData = None, parent=None):
        QtCore.QAbstractTableModel.__init__(self, parent)
        if table_data is not None:
            self.init_data(table_data)

    def init_data(self, table_data: PandasTableData):
        self.table_data = table_data

    def flags(self, index):
        flags = super().flags(index)
        protected_cell_ranges = self.table_data.protected_cell_ranges
        cell = string_from_index((index.row(), index.column()))
        is_contain = [protected_range for protected_range in protected_cell_ranges if cell in protected_range]
        if len(is_contain) > 0:
            # if string_from_index((index.row(), index.column())) in protected_cell_ranges:
            flags = flags & ~Qt.ItemIsEditable
        else:
            flags = flags | Qt.ItemIsEditable
        return flags

    def rowCount(self, parent=QtCore.QModelIndex()):
        return len(self.table_data.dataframe.index)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return self.table_data.dataframe.columns.size

    def setData(self, index: QtCore.QModelIndex, value: QtCore.QVariant, role=Qt.EditRole):
        if role == Qt.EditRole:
            row = index.row()
            column = index.column()

            def setData2(row_, column_):
                # update dataframe
                self.table_data.dataframe.iloc[row_, column_] = value

            setData2(row, column)

            cell_letter = self.table_data.index_to_cell_letter(row, column)
            if self.table_data.is_cell_range_contain(cell_letter, self.table_data.merged_cell_ranges):
                cell_range = None
                for cell_range_ in self.table_data.merged_cell_ranges:
                    if cell_letter in cell_range_:
                        cell_range = cell_range_
                        break
                for row in range(cell_range.min_row - 1, cell_range.max_row):
                    for column in range(cell_range.min_col - 1, cell_range.max_col):
                        setData2(row, column)
            else:
                setData2(row, column)

        return True

    def data(self, index, role=Qt.DisplayRole):
        row = index.row()
        column = index.column()
        if False and index.column() == temp_column_number:
            # if role == Qt.CheckStateRole:
            #     return True
            if role == Qt.DisplayRole:
                value = self.table_data.dataframe.iloc[index.row(), index.column()]
                return value
            if role == Qt.EditRole:
                value = self.table_data.dataframe.iloc[index.row(), index.column()]
                return value
        else:
            if role == Qt.DisplayRole:
                return QtCore.QVariant('{0}'.format(self.table_data.dataframe.iloc[row, column]))
            elif role == Qt.EditRole:
                return QtCore.QVariant('{0}'.format(self.table_data.dataframe.iloc[row, column]))
        return QtCore.QVariant()

        # if index.isValid():
        #     if role == Qt.DisplayRole or role == Qt.EditRole:
        #         string = str(self.table_data.dataframe.iloc[index.row()][index.column()])
        #         return QtCore.QVariant(string)
        # return QtCore.QVariant()

    def headerData(self, section: int, orientation: Qt.Orientation, role: Qt.ItemDataRole = Qt.DisplayRole):
        """Override method from QAbstractTableModel

        Return dataframe index as vertical header data and columns as horizontal header data.
        """
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                header_item = self.table_data.get_shown_column(section)
                if isinstance(header_item, tuple):
                    raise Exception("Dataframe 의 헤더가 여러줄일 수 없습니다.")
                else:
                    header_text = header_item
                return header_text
            if orientation == Qt.Vertical:
                index_str = self.table_data.get_shown_row(section)
                return index_str
        return None


class CheckBoxDelegate(QtWidgets.QStyledItemDelegate):
    """
    A delegate that places a fully functioning QCheckBox in every
    cell of the column to which it's applied
    """
    TRUE_KEY = 'Y'
    FALSE_KEY = 'N'

    def __init__(self, parent):
        super().__init__(parent)

    def to_bool(self, str_):
        if str_ == self.TRUE_KEY:
            return True
        elif str_ == self.FALSE_KEY:
            return False
        else:
            return False

    def toggle_bool_str(self, str_):
        if str_ == self.TRUE_KEY:
            return self.FALSE_KEY
        elif str_ == self.FALSE_KEY:
            return self.TRUE_KEY
        else:
            return self.TRUE_KEY

    def createEditor(self, parent, option, index):
        """
        Important, otherwise an editor is created if the user clicks in this cell.
        ** Need to hook up a signal to the model
        """
        return None

    def paint(self, painter, option, index):
        """
        Paint a checkbox without the label.
        """

        checked = self.to_bool(index.data())  # .toBool()
        # checked = str(index.data())
        check_box_style_option = QtWidgets.QStyleOptionButton()

        if Qt.ItemIsEditable & index.flags().__index__() > 0:
            check_box_style_option.state |= QtWidgets.QStyle.State_Enabled
        else:
            check_box_style_option.state |= QtWidgets.QStyle.State_ReadOnly

        if checked:
            check_box_style_option.state |= QtWidgets.QStyle.State_On
        else:
            check_box_style_option.state |= QtWidgets.QStyle.State_Off

        check_box_style_option.rect = self.getCheckBoxRect(option)

        check_box_style_option.state |= QtWidgets.QStyle.State_Enabled

        QtWidgets.QApplication.style().drawControl(QtWidgets.QStyle.CE_CheckBox, check_box_style_option, painter)

    def editorEvent(self, event, model, option, index):
        """
        Change the data in the model and the state of the checkbox
        if the user presses the left mousebutton or presses
        Key_Space or Key_Select and this cell is editable. Otherwise do nothing.
        """
        # print('Check Box editor Event detected : ')
        if not (Qt.ItemIsEditable & index.flags().__index__()) > 0:
            return False

        # print('Check Box editor Event detected : passed first check')
        # Do not change the checkbox-state
        if event.type() == QEvent.MouseButtonPress:
            return False
        if event.type() == QEvent.MouseButtonRelease or event.type() == QEvent.MouseButtonDblClick:
            if event.button() != Qt.LeftButton or not self.getCheckBoxRect(option).contains(event.pos()):
                return False
            if event.type() == QEvent.MouseButtonDblClick:
                return True
        elif event.type() == QEvent.KeyPress:
            if event.key() != Qt.Key_Space and event.key() != Qt.Key_Select:
                return False
        else:
            return False

        # Change the checkbox-state
        self.setModelData(None, model, index)
        return True

    def setModelData(self, editor, model, index):
        """
        The user wanted to change the old state in the opposite.
        """
        # print('SetModelData')
        newValue = self.toggle_bool_str(index.data())
        # print('New Value : {0}'.format(newValue))
        model.setData(index, newValue, Qt.EditRole)

    @staticmethod
    def getCheckBoxRect(option):
        check_box_style_option = QtWidgets.QStyleOptionButton()
        check_box_rect = QtWidgets.QApplication.style().subElementRect(QtWidgets.QStyle.SE_CheckBoxIndicator,
                                                                       check_box_style_option, None)
        check_box_point = QtCore.QPoint(option.rect.x() +
                                        option.rect.width() / 2 -
                                        check_box_rect.width() / 2,
                                        option.rect.y() +
                                        option.rect.height() / 2 -
                                        check_box_rect.height() / 2)
        return QtCore.QRect(check_box_point, check_box_rect.size())


class ComboBoxDelegate(QtWidgets.QStyledItemDelegate):
    """
    A delegate that places a fully functioning QCheckBox in every
    cell of the column to which it's applied
    """
    TRUE_KEY = 'Y'
    FALSE_KEY = 'N'

    def __init__(self, parent, item_text_list: List[str]):
        super().__init__(parent)
        self.item_text_list = item_text_list

    def createEditor(self, parent, option, index):
        """
        Important, otherwise an editor is created if the user clicks in this cell.
        ** Need to hook up a signal to the model
        """
        return None

    def paint(self, painter, option, index):
        """
        Paint a checkbox without the label.
        """

        value = index.data()  # .toBool()
        # checked = str(index.data())
        combo_box = QtWidgets.QComboBox()
        for item_text in self.item_text_list:
            combo_box.addItem(item_text)

        if Qt.ItemIsEditable & index.flags().__index__() > 0:
            combo_box.setEditable(False)
        else:
            combo_box.setEditable(True)

        if value in self.item_text_list:
            combo_box.setCurrentText(value)
        else:
            combo_box.clear()

        # combo_box.rect = self.getCheckBoxRect(option)

        # combo_box.state |= QtWidgets.QStyle.State_Enabled

        # QtWidgets.QApplication.style().drawControl(QtWidgets.QStyle.CE_CheckBox, combo_box, painter)

    def editorEvent(self, event, model, option, index):
        """
        Change the data in the model and the state of the checkbox
        if the user presses the left mousebutton or presses
        Key_Space or Key_Select and this cell is editable. Otherwise do nothing.
        """
        # print('Check Box editor Event detected : ')
        if not (Qt.ItemIsEditable & index.flags().__index__()) > 0:
            return False

        # print('Check Box editor Event detected : passed first check')
        # Do not change the checkbox-state
        if event.type() == QEvent.MouseButtonPress:
            return False
        if event.type() == QEvent.MouseButtonRelease or event.type() == QEvent.MouseButtonDblClick:
            if event.button() != Qt.LeftButton or not self.getCheckBoxRect(option).contains(event.pos()):
                return False
            if event.type() == QEvent.MouseButtonDblClick:
                return True
        elif event.type() == QEvent.KeyPress:
            if event.key() != Qt.Key_Space and event.key() != Qt.Key_Select:
                return False
        else:
            return False

        # Change the checkbox-state
        self.setModelData(None, model, index)
        return True

    def setModelData(self, editor, model, index):
        """
        The user wanted to change the old state in the opposite.
        """
        # print('SetModelData')
        a = index.data()
        newValue = self.toggle_bool_str(index.data())
        # print('New Value : {0}'.format(newValue))
        model.setData(index, newValue, Qt.EditRole)


class ComboDelegate(QtWidgets.QItemDelegate):
    height = 25
    width = 200

    def __init__(self, parent, item_text_list: List[str]):
        super().__init__(parent)
        self.item_text_list = [''] + item_text_list

    def createEditor(self, parent, option, index):
        editor = QtWidgets.QComboBox(parent)
        editor.addItems(self.item_text_list)
        # editor.addItems(self.editorItems)
        # editor.setEditable(True)
        editor.currentIndexChanged.connect(self.currentIndexChanged)

        return editor

    def setEditorData(self, editor, index):
        if index.data() in self.item_text_list:
            editor.setCurrentText(index.data())
        else:
            editor.setCurrentText('')
        # editor.setGeometry(0, index.row()*self.height, self.width, self.height*len(self.item_text_list))
        editor.showPopup()

    def setModelData(self, editor, model, index):
        editorIndex = editor.currentIndex()
        text = editor.currentText()
        model.setData(index, text)
        # print '\t\t\t ...setModelData() 1', text

    @QtCore.pyqtSlot()
    def currentIndexChanged(self):
        self.commitData.emit(self.sender())
